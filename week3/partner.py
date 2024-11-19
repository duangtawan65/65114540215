#import pandas as pd

# Load the Excel file
#df = pd.read_excel('Me & My partner.xlsx')


#filtered_df = df.dropna(how='all', subset=['My Github User', "My partner's github user"])
#print(filtered_df)

from openpyxl import load_workbook
wb = load_workbook('Me & My partner.xlsx')
sheet = wb.active

groups = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    values = list(row)

    if values[1] is not None and values[2] is not None:
        pair = (values[1], values[2])
        if values[1] < values[2]:
            pair = (values[2], values[1])
        if pair not in groups:
            groups.append(pair)
            print(pair)


